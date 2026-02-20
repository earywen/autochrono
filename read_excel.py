import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\l.brigaud\Desktop\Vibecoding\autochrono\Chrono 2026.xlsx')
ws = wb.active

print("=== STRUCTURE DU FICHIER CHRONO 2026 ===\n")
print("COLONNES (ligne 1):")
for i, cell in enumerate(ws[1], 1):
    val = str(cell.value).replace('\n', ' ')[:60] if cell.value else "(vide)"
    print(f"  {cell.column_letter}: {val}")

print("\n=== EXEMPLE DONNEES (ligne 6) ===")
row6 = list(ws[6])
headers = list(ws[1])
for i, (header, cell) in enumerate(zip(headers, row6)):
    h = str(header.value).replace('\n', ' ')[:20] if header.value else f"Col{i+1}"
    v = str(cell.value)[:50] if cell.value else "(vide)"
    print(f"  {h}: {v}")

print("\n=== DERNIERE LIGNE AVEC DONNEES ===")
last_row = ws.max_row
for r in range(ws.max_row, 1, -1):
    if ws.cell(r, 1).value:
        last_row = r
        break

print(f"Ligne {last_row}:")
for i, (header, cell) in enumerate(zip(headers, ws[last_row])):
    h = str(header.value).replace('\n', ' ')[:20] if header.value else f"Col{i+1}"
    v = str(cell.value)[:50] if cell.value else "(vide)"
    print(f"  {h}: {v}")
